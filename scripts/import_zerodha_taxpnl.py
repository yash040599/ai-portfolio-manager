"""
Import and verify trade data from a Zerodha Tax P&L xlsx report.

• Intraday section  → verifies / corrects existing rows in intraday_tax_ledger,
                      inserts missing trades, marks everything as 'verified'.
• Short-term / Long-term sections → inserts into capital_gains_ledger.
• Also updates trading_data_*.json and trading_report_*.txt with corrected
  values, adds a ✓ VERIFIED header and an "Updated on …" remark.

Usage
─────
    python scripts/import_zerodha_taxpnl.py                              # latest xlsx in data/ZerodhaTaxPL/
    python scripts/import_zerodha_taxpnl.py data/ZerodhaTaxPL/file.xlsx  # specific file
"""

import argparse
import datetime
import glob
import json
import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from scripts.tax_db import get_db, indian_fy, fy_label
from config import now_ist

ZERODHA_DIR = os.path.join(PROJECT_ROOT, "data", "ZerodhaTaxPL")


# ── xlsx parsing ──────────────────────────────────────────────────

def _to_date_str(val) -> str:
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.isoformat()
    return str(val).strip() if val else ""


def _f(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def parse_xlsx(path: str):
    """
    Parse the 'Tradewise Exits' sheet.
    Returns (intraday_trades, short_term_trades, long_term_trades).
    Each trade is a dict with Zerodha fields.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # First sheet is always Tradewise Exits
    ws = wb.worksheets[0]

    intraday, short_term, long_term = [], [], []
    section = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        cell1 = vals[1] if len(vals) > 1 else None

        # Section headers
        if cell1 == "Equity - Intraday":
            section = "intraday"; continue
        if cell1 == "Equity - Short Term":
            section = "short_term"; continue
        if cell1 == "Equity - Long Term":
            section = "long_term"; continue
        if cell1 in ("Equity - Buyback", "Non Equity", "Mutual Funds",
                      "F&O", "Currency", "Commodity"):
            section = None; continue
        if cell1 == "Symbol":      # header row — skip
            continue

        if section is None or cell1 is None or not str(cell1).strip():
            continue

        t = {
            "symbol":       str(cell1).strip(),
            "isin":         str(vals[2]).strip() if vals[2] else "",
            "entry_date":   _to_date_str(vals[3]),
            "exit_date":    _to_date_str(vals[4]),
            "qty":          _f(vals[5]),
            "buy_value":    _f(vals[6]),
            "sell_value":   _f(vals[7]),
            "profit":       _f(vals[8]),
            "holding_days": int(_f(vals[9])),
            "fmv":          _f(vals[10]),
            "taxable_profit": _f(vals[11]),
            "turnover":     _f(vals[12]),
            "brokerage":    _f(vals[13]),
            "exchange_txn": _f(vals[14]) + _f(vals[15]),   # exch + IPFT
            "sebi_charges": _f(vals[16]),
            "gst":          _f(vals[17]) + _f(vals[18]) + _f(vals[19]),  # CGST+SGST+IGST
            "stamp_duty":   _f(vals[20]),
            "stt":          _f(vals[21]),
        }
        t["total_charges"] = round(
            t["brokerage"] + t["exchange_txn"] + t["sebi_charges"]
            + t["gst"] + t["stamp_duty"] + t["stt"], 4
        )

        if section == "intraday":
            intraday.append(t)
        elif section == "short_term":
            short_term.append(t)
        elif section == "long_term":
            long_term.append(t)

    wb.close()
    return intraday, short_term, long_term


# ── Intraday verification ────────────────────────────────────────

def _verify_intraday(conn, zerodha_trades: list[dict]) -> dict:
    """
    Verify / correct intraday_tax_ledger against Zerodha data.
    Groups by (date, symbol) and compares aggregate P&L.

    Returns {"verified": n, "corrected": n, "inserted": n}.
    """
    # Group Zerodha by (exit_date, symbol)
    z_groups: dict[tuple, list] = {}
    for t in zerodha_trades:
        key = (t["exit_date"], t["symbol"])
        z_groups.setdefault(key, []).append(t)

    # Group DB by (date, symbol)
    db_rows = conn.execute(
        "SELECT * FROM intraday_tax_ledger"
    ).fetchall()
    db_groups: dict[tuple, list] = {}
    for r in db_rows:
        key = (r["date"], r["symbol"])
        db_groups.setdefault(key, []).append(r)

    stats = {"verified": 0, "corrected": 0, "inserted": 0}

    for key, z_trades in z_groups.items():
        date, symbol = key
        z_total_pnl = sum(t["profit"] for t in z_trades)
        z_total_qty = sum(t["qty"] for t in z_trades)

        if key in db_groups:
            db_rows_g = db_groups[key]
            db_total_pnl = sum(r["gross_pnl"] for r in db_rows_g)
            db_total_qty = sum(r["qty"] for r in db_rows_g)

            if (abs(db_total_pnl - z_total_pnl) < 0.10
                    and abs(db_total_qty - z_total_qty) < 0.01):
                # Match — mark existing rows verified
                for r in db_rows_g:
                    conn.execute(
                        "UPDATE intraday_tax_ledger SET verified='verified' WHERE id=?",
                        (r["id"],),
                    )
                stats["verified"] += len(db_rows_g)
            else:
                # Mismatch — replace with Zerodha data
                for r in db_rows_g:
                    conn.execute(
                        "DELETE FROM intraday_tax_ledger WHERE id=?",
                        (r["id"],),
                    )
                for i, t in enumerate(z_trades):
                    _insert_zerodha_intraday(conn, t, i)
                stats["corrected"] += len(z_trades)
                print(f"    ✎ Corrected {symbol} on {date}: "
                      f"P&L {db_total_pnl:+.2f} → {z_total_pnl:+.2f}")
        else:
            # Only in Zerodha — insert
            for i, t in enumerate(z_trades):
                _insert_zerodha_intraday(conn, t, i)
            stats["inserted"] += len(z_trades)

    conn.commit()

    # ── Fix sides and entry/exit for all rows ─────────────────
    # Ensure side matches the trades table and entry/exit are set
    # from the correct perspective using the stored buy_value/sell_value.
    fixed = 0
    all_ledger = conn.execute(
        "SELECT id, date, symbol, side, qty, entry_price, exit_price, "
        "       buy_value, sell_value "
        "FROM intraday_tax_ledger"
    ).fetchall()

    # Build a lookup of trade sides: (date, symbol, qty) → side
    all_trades = conn.execute(
        "SELECT symbol, side, qty, date FROM trades"
    ).fetchall()
    trade_sides: dict[tuple, str] = {}
    for tr in all_trades:
        key = (tr["date"], tr["symbol"], tr["qty"])
        trade_sides[key] = tr["side"]

    for row in all_ledger:
        key = (row["date"], row["symbol"], row["qty"])
        actual_side = trade_sides.get(key, "BUY")

        qty = row["qty"]
        if qty == 0:
            continue

        # Recalculate entry/exit from stored buy/sell values
        buy_per_unit  = round(row["buy_value"] / qty, 2)
        sell_per_unit = round(row["sell_value"] / qty, 2)

        if actual_side == "SELL":
            correct_entry = sell_per_unit   # we entered by selling
            correct_exit  = buy_per_unit    # we exited by buying back
        else:
            correct_entry = buy_per_unit    # we entered by buying
            correct_exit  = sell_per_unit   # we exited by selling

        if (row["side"] != actual_side
                or abs(row["entry_price"] - correct_entry) > 0.001
                or abs(row["exit_price"] - correct_exit) > 0.001):
            conn.execute(
                "UPDATE intraday_tax_ledger "
                "SET side=?, entry_price=?, exit_price=? WHERE id=?",
                (actual_side, correct_entry, correct_exit, row["id"]),
            )
            fixed += 1

    if fixed:
        conn.commit()
        stats["sides_fixed"] = fixed

    return stats


def _insert_zerodha_intraday(conn, t: dict, idx: int):
    """Insert a single Zerodha intraday trade into the DB.

    Cross-references the trades table to determine original side (BUY/SELL)
    and sets entry/exit from the bot's perspective accordingly.
    """
    date = t["exit_date"]
    qty = int(t["qty"])
    buy_price  = round(t["buy_value"] / qty, 2) if qty else 0
    sell_price = round(t["sell_value"] / qty, 2) if qty else 0
    order_id = f"ZV_{date}_{t['symbol']}_{idx}"

    # Look up original side from trades table
    side = "BUY"  # default
    entry_price, exit_price = buy_price, sell_price

    db_rows = conn.execute(
        "SELECT side, entry_price FROM trades WHERE date=? AND symbol=? AND qty=?",
        (date, t["symbol"], qty),
    ).fetchall()
    for row in db_rows:
        if row["side"] == "SELL":
            side = "SELL"
            entry_price = sell_price   # we entered by selling
            exit_price  = buy_price    # we exited by buying back
            break

    conn.execute(
        """INSERT OR REPLACE INTO intraday_tax_ledger
           (date, symbol, exchange, side, qty,
            entry_price, exit_price, entry_time, exit_time,
            exit_reason, gross_pnl,
            buy_value, sell_value, turnover,
            brokerage, stt, exchange_txn, gst,
            sebi_charges, stamp_duty, total_charges,
            net_pnl, order_id, verified)
           VALUES (?,?,?,?,?, ?,?,?,?, ?,?, ?,?,?, ?,?,?,?, ?,?,?, ?,?,?)""",
        (
            date, t["symbol"], "NSE", side, qty,
            entry_price, exit_price, "", "",
            "", round(t["profit"], 2),
            round(t["buy_value"], 2), round(t["sell_value"], 2),
            round(t["turnover"], 2),
            round(t["brokerage"], 4), round(t["stt"], 4),
            round(t["exchange_txn"], 4), round(t["gst"], 4),
            round(t["sebi_charges"], 4), round(t["stamp_duty"], 4),
            round(t["total_charges"], 4),
            round(t["profit"] - t["total_charges"], 2),
            order_id, "verified",
        ),
    )


# ── Update trading reports (JSON + TXT) ──────────────────────────

def _trading_file_paths(date: datetime.date) -> tuple[str, str]:
    """Return (json_path, txt_path) for a trading date."""
    d = date
    base = os.path.join(
        PROJECT_ROOT, "reports", "trading",
        str(d.year), f"{d.month:02d}",
    )
    json_path = os.path.join(base, f"trading_data_{d.day:02d}.json")
    txt_path  = os.path.join(base, f"trading_report_{d.day:02d}.txt")
    return json_path, txt_path


def _match_positions_from_trades_db(conn, positions: list[dict], date_str: str):
    """
    Match JSON positions to the trades DB and update entry_price,
    exit_price, pnl from the reconciled DB values.
    The trades DB is the authoritative source (already reconciled
    with Zerodha live data).
    """
    db_rows = conn.execute(
        "SELECT symbol, side, qty, entry_price, exit_price, pnl "
        "FROM trades WHERE date=?",
        (date_str,),
    ).fetchall()

    # Build a list of unmatched DB rows
    unmatched_db = [dict(r) for r in db_rows]

    for pos in positions:
        if pos.get("status") != "CLOSED":
            continue

        symbol = pos["symbol"]
        side   = pos.get("side", "BUY")
        qty    = pos["qty"]

        # Find best matching DB row by (symbol, side, qty, closest entry)
        best = None
        best_diff = float("inf")
        for db_row in unmatched_db:
            if (db_row["symbol"] != symbol
                    or db_row["side"] != side
                    or db_row["qty"] != qty):
                continue
            diff = abs(db_row["entry_price"] - pos["entry_price"])
            if diff < best_diff:
                best = db_row
                best_diff = diff

        if best:
            unmatched_db.remove(best)
            pos["entry_price"] = best["entry_price"]
            pos["exit_price"]  = best["exit_price"]
            pos["pnl"]         = best["pnl"]


def _aggregate_zerodha_charges(z_trades: list[dict]) -> dict:
    """Sum Zerodha's actual charges across all trades for a date."""
    return {
        "brokerage":     round(sum(t["brokerage"]    for t in z_trades), 4),
        "stt":           round(sum(t["stt"]          for t in z_trades), 4),
        "exchange_txn":  round(sum(t["exchange_txn"] for t in z_trades), 4),
        "gst":           round(sum(t["gst"]          for t in z_trades), 4),
        "sebi_charges":  round(sum(t["sebi_charges"] for t in z_trades), 4),
        "stamp_duty":    round(sum(t["stamp_duty"]   for t in z_trades), 4),
    }


def _update_trading_reports(zerodha_trades: list[dict], conn) -> dict:
    """
    Update trading_data_*.json and trading_report_*.txt files.

    - Position values (entry/exit/pnl) come from the trades DB
      (already reconciled with Zerodha live data).
    - Charges come from Zerodha Tax P&L actuals.

    Returns {"updated": [dates], "skipped": [dates]}.
    """
    # Group Zerodha trades by exit_date
    by_date: dict[str, list] = {}
    for t in zerodha_trades:
        by_date.setdefault(t["exit_date"], []).append(t)

    stats: dict[str, list] = {"updated": [], "skipped": []}

    for date_str, z_trades in sorted(by_date.items()):
        d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        json_path, txt_path = _trading_file_paths(d)

        if not os.path.exists(json_path):
            stats["skipped"].append(date_str)
            continue

        # ── Load existing JSON ────────────────────────────────
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        positions = data.get("positions", [])
        old_charges = data.get("pnl", {}).get("charges", {})

        # ── Update positions from trades DB ───────────────────
        _match_positions_from_trades_db(conn, positions, date_str)

        # ── Recalculate P&L with actual Zerodha charges ──────
        z_charges = _aggregate_zerodha_charges(z_trades)
        total_buy = round(sum(t["buy_value"] for t in z_trades), 2)
        total_sell = round(sum(t["sell_value"] for t in z_trades), 2)
        total_turnover = round(total_buy + total_sell, 2)
        num_orders = old_charges.get("num_orders", len(z_trades) * 2)

        total_tax_and_charges = round(
            z_charges["brokerage"] + z_charges["stt"] +
            z_charges["exchange_txn"] + z_charges["gst"] +
            z_charges["sebi_charges"] + z_charges["stamp_duty"], 2
        )

        # Preserve Claude API cost from original report
        claude_api_cost = old_charges.get("claude_api_cost", 0.0)
        total_costs = round(total_tax_and_charges + claude_api_cost, 2)

        gross_pnl = round(
            sum(p.get("pnl", 0) for p in positions if p.get("status") == "CLOSED"), 2
        )
        net_profit = round(gross_pnl - total_costs, 2)

        # Tax estimate
        from config import Config
        tax_rate = Config.TAX_RATE_PCT * (1 + Config.TAX_CESS_PCT / 100) / 100
        tax_rate_pct = round(Config.TAX_RATE_PCT * (1 + Config.TAX_CESS_PCT / 100), 2)
        estimated_tax = round(net_profit * tax_rate, 2) if net_profit > 0 else 0.0
        profit_after_tax = round(net_profit - estimated_tax, 2)

        charges = {
            "total_turnover":        total_turnover,
            "buy_turnover":          total_buy,
            "sell_turnover":         total_sell,
            "num_orders":            num_orders,
            "brokerage":             round(z_charges["brokerage"], 2),
            "stt":                   round(z_charges["stt"], 2),
            "exchange_txn":          round(z_charges["exchange_txn"], 2),
            "gst":                   round(z_charges["gst"], 2),
            "sebi_charges":          z_charges["sebi_charges"],
            "stamp_duty":            round(z_charges["stamp_duty"], 2),
            "total_tax_and_charges": total_tax_and_charges,
            "claude_api_cost":       claude_api_cost,
            "total_costs":           total_costs,
            "zerodha_monthly_fyi":   old_charges.get("zerodha_monthly_fyi", 500.0),
        }

        pnl = {
            "gross_pnl":        gross_pnl,
            "charges":          charges,
            "net_profit":       net_profit,
            "is_profitable":    net_profit > 0,
            "tax_rate_pct":     tax_rate_pct,
            "estimated_tax":    estimated_tax,
            "profit_after_tax": profit_after_tax,
        }

        # ── Write updated JSON ────────────────────────────────
        now_str = now_ist().strftime("%Y-%m-%d %H:%M:%S")
        data["positions"] = positions
        data["pnl"] = pnl
        data["verified"] = True
        data["verified_on"] = now_str

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)

        # ── Regenerate TXT report ─────────────────────────────
        _write_verified_txt(
            txt_path, data, positions, charges, pnl, now_str
        )

        stats["updated"].append(date_str)
        print(f"    ✓ Updated reports for {date_str}")

    return stats


def _write_verified_txt(
    txt_path: str,
    data: dict,
    positions: list[dict],
    charges: dict,
    pnl: dict,
    verified_on: str,
):
    """Regenerate the .txt trading report from corrected data."""
    SEP_MAJOR = "=" * 58
    SEP_MINOR = "─" * 58
    SEP_TABLE = "─" * 86

    date_str = data["date"]
    mode_label = "DRY RUN (simulated)" if data.get("mode") == "dry_run" else "LIVE TRADING"
    session_count = data.get("sessions", 1)
    config = data.get("config", {})
    budget = config.get("budget", 0)
    market_condition = data.get("market_condition", "")
    trade_log = data.get("trade_log", [])

    from config import Config

    closed = [p for p in positions if p.get("status") == "CLOSED"]
    open_p = [p for p in positions if p.get("status") == "OPEN"]
    winners = [p for p in closed if p.get("pnl", 0) > 0]
    losers  = [p for p in closed if p.get("pnl", 0) < 0]

    with open(txt_path, "w", encoding="utf-8") as f:
        # ── Verified header ───────────────────────────────────
        f.write(f"{'=' * 58}\n")
        f.write(f"  ✓ VERIFIED — Data corrected from Zerodha Tax P&L\n")
        f.write(f"  Updated on: {verified_on}\n")
        f.write(f"{'=' * 58}\n\n")

        # ── Header ────────────────────────────────────────────
        f.write(f"{SEP_MAJOR}\n")
        f.write(f"  INTRADAY TRADING REPORT — {date_str}\n")
        f.write(f"  Mode: {mode_label}\n")
        if session_count > 1:
            f.write(f"  Sessions: {session_count} (combined)\n")
        f.write(f"{SEP_MAJOR}\n\n")

        # ── Configuration ─────────────────────────────────────
        f.write("CONFIGURATION\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"Claude plan     : {config.get('claude_plan', 'PRO').upper()}\n")
        f.write(f"Budget          : ₹{budget:,.2f} (from Zerodha funds)\n")
        f.write(f"Universe        : {config.get('universe', 'NIFTY100')}\n")
        if market_condition:
            f.write(f"Market condition: {market_condition}\n")
        f.write(f"Max positions   : {config.get('max_positions', 5)}\n")
        f.write(f"Stop-loss       : {config.get('stop_loss_pct', 1.5)}%\n")
        f.write(f"Target          : {config.get('target_pct', 2.0)}%\n")
        f.write(f"Circuit breaker : {Config.MAX_LOSS_PER_DAY_PCT}%\n\n")

        # ── Trade Summary ─────────────────────────────────────
        f.write("TRADE SUMMARY\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"Total trades    : {len(closed)}\n")
        f.write(f"Winners         : {len(winners)}\n")
        f.write(f"Losers          : {len(losers)}\n")
        f.write(f"Still open      : {len(open_p)}\n")
        if closed:
            win_rate = len(winners) / len(closed) * 100
            f.write(f"Win rate        : {win_rate:.1f}%\n")
        f.write("\n")

        # ── Trade Details Table ───────────────────────────────
        f.write("TRADE DETAILS\n")
        f.write(f"{SEP_TABLE}\n")
        f.write(
            f"{'SYMBOL':<12} {'SIDE':<6} {'QTY':>5} "
            f"{'ENTRY':>10} {'EXIT':>10} {'P&L':>12} "
            f"{'REASON':<14} {'ENTRY_T':<10} {'EXIT_T':<10}\n"
        )
        f.write(f"{SEP_TABLE}\n")

        for p in positions:
            exit_p  = f"₹{p['exit_price']:.2f}" if p.get("exit_price") else "—"
            pnl_val = f"₹{p.get('pnl', 0):+,.2f}" if p.get("exit_price") else "—"
            f.write(
                f"{p['symbol']:<12} {p['side']:<6} {p['qty']:>5} "
                f"₹{p['entry_price']:>9.2f} {exit_p:>10} {pnl_val:>12} "
                f"{(p.get('exit_reason') or 'OPEN'):<14} "
                f"{(p.get('entry_time') or '—'):<10} "
                f"{(p.get('exit_time') or '—'):<10}\n"
            )

        f.write("\n")

        # ── Rationales ────────────────────────────────────────
        f.write("TRADE RATIONALES\n")
        f.write(f"{SEP_MINOR}\n")
        for p in positions:
            f.write(f"  {p['symbol']}: {p.get('rationale', '—')}\n")
        f.write("\n")

        # ── P&L Breakdown ─────────────────────────────────────
        f.write(f"{SEP_MAJOR}\n")
        f.write("P&L BREAKDOWN\n")
        f.write(f"{SEP_MAJOR}\n\n")

        f.write(f"Gross P&L               : ₹{pnl['gross_pnl']:+,.2f}\n\n")

        f.write("CHARGES & TAXES:\n")
        f.write(f"  Brokerage             : ₹{charges['brokerage']:,.2f}\n")
        f.write(f"  STT (sell side)       : ₹{charges['stt']:,.2f}\n")
        f.write(f"  Exchange transaction  : ₹{charges['exchange_txn']:,.2f}\n")
        f.write(f"  GST (18%)             : ₹{charges['gst']:,.2f}\n")
        f.write(f"  SEBI charges          : ₹{charges['sebi_charges']:,.4f}\n")
        f.write(f"  Stamp duty (buy side) : ₹{charges['stamp_duty']:,.2f}\n")
        f.write(f"  {'─' * 40}\n")
        f.write(f"  Total tax & charges   : ₹{charges['total_tax_and_charges']:,.2f}\n\n")

        f.write("CLAUDE API COST:\n")
        f.write(f"  Claude API usage      : ₹{charges['claude_api_cost']:,.2f}  (est. ₹{Config.CLAUDE_COST_PER_CALL}/call × actual calls)\n")
        f.write(f"  {'─' * 40}\n")
        f.write(f"  Total all costs       : ₹{charges['total_costs']:,.2f}\n\n")

        f.write(f"{'=' * 42}\n")
        f.write(f"  NET PROFIT AFTER ALL  : ₹{pnl['net_profit']:+,.2f}\n")
        f.write(f"{'=' * 42}\n")
        profitable = "YES ✓" if pnl["is_profitable"] else "NO ✗"
        f.write(f"  Profitable?           : {profitable}\n")
        if budget > 0:
            returns_pct = pnl["net_profit"] / budget * 100
            f.write(f"  Day returns           : {returns_pct:+.2f}% on ₹{budget:,.0f} budget\n")
        f.write("\n")

        # ── Estimated Income Tax ──────────────────────────────
        tax_rate_pct = pnl.get("tax_rate_pct", 0)
        estimated_tax = pnl.get("estimated_tax", 0)
        f.write("ESTIMATED INCOME TAX (speculative business income)\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"  Tax slab rate         : {Config.TAX_RATE_PCT}% + {Config.TAX_CESS_PCT}% cess = {tax_rate_pct}% effective\n")
        if pnl["net_profit"] > 0:
            f.write(f"  Estimated tax         : ₹{estimated_tax:,.2f}\n")
            f.write(f"  Profit after tax      : ₹{pnl['profit_after_tax']:+,.2f}\n")
        else:
            f.write(f"  Estimated tax         : ₹0.00 (no tax on losses)\n")
            f.write(f"  Loss can be carried forward for 4 years (speculative only)\n")
        f.write("\n")

        f.write(f"  FYI: Zerodha Kite Connect subscription is ₹{Config.ZERODHA_MONTHLY_COST:,.0f}/month (not deducted above).\n")
        f.write(f"  Track cumulative daily profits to ensure they cover this monthly cost.\n\n")

        # ── Turnover Details ──────────────────────────────────
        f.write("TURNOVER DETAILS\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"  Buy turnover          : ₹{charges['buy_turnover']:,.2f}\n")
        f.write(f"  Sell turnover         : ₹{charges['sell_turnover']:,.2f}\n")
        f.write(f"  Total turnover        : ₹{charges['total_turnover']:,.2f}\n")
        f.write(f"  Total orders          : {charges['num_orders']}\n\n")

        # ── Chronological Trade Log ───────────────────────────
        if trade_log:
            f.write("CHRONOLOGICAL TRADE LOG\n")
            f.write(f"{SEP_MINOR}\n")
            for entry in trade_log:
                f.write(
                    f"  [{entry['time']}] {entry['action']:<14} "
                    f"{entry['symbol']:<12} {entry['side']:<5} "
                    f"{entry['qty']:>5}  ₹{entry['price']:>10}  "
                    f"{entry['detail']}\n"
                )
            f.write("\n")


# ── Capital gains import ─────────────────────────────────────────

def _import_capital_gains(conn, trades: list[dict], trade_type: str) -> int:
    """Insert capital-gains trades. Skips duplicates. Returns insert count."""
    inserted = 0
    for t in trades:
        qty = t["qty"]
        try:
            conn.execute(
                """INSERT INTO capital_gains_ledger
                   (trade_type, symbol, isin, entry_date, exit_date,
                    qty, buy_value, sell_value, profit,
                    period_of_holding, fair_market_value, taxable_profit,
                    turnover, brokerage, exchange_txn, sebi_charges,
                    gst, stamp_duty, stt, total_charges, verified)
                   VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?, ?,?,?,?, ?,?,?,?,?)""",
                (
                    trade_type, t["symbol"], t["isin"],
                    t["entry_date"], t["exit_date"],
                    qty, round(t["buy_value"], 2), round(t["sell_value"], 2),
                    round(t["profit"], 2),
                    t["holding_days"], round(t["fmv"], 2),
                    round(t["taxable_profit"], 2),
                    round(t["turnover"], 2),
                    round(t["brokerage"], 4), round(t["exchange_txn"], 4),
                    round(t["sebi_charges"], 4), round(t["gst"], 4),
                    round(t["stamp_duty"], 4), round(t["stt"], 4),
                    round(t["total_charges"], 4), "verified",
                ),
            )
            inserted += 1
        except Exception:
            pass  # duplicate — silently skip
    conn.commit()
    return inserted


# ── CLI ───────────────────────────────────────────────────────────

def _find_latest_xlsx() -> str | None:
    pattern = os.path.join(ZERODHA_DIR, "taxpnl-*.xlsx")
    files = sorted(glob.glob(pattern))
    return files[-1] if files else None


def main():
    parser = argparse.ArgumentParser(
        description="Import Zerodha Tax P&L xlsx — verify intraday & import capital gains.",
    )
    parser.add_argument(
        "xlsx", nargs="?", default=None,
        help="Path to Zerodha xlsx. Default: latest file in data/ZerodhaTaxPL/.",
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx or _find_latest_xlsx()
    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f"\n  No Zerodha xlsx found. Place it in data/ZerodhaTaxPL/")
        return

    print(f"\n  Reading: {os.path.basename(xlsx_path)}")
    intraday, short_term, long_term = parse_xlsx(xlsx_path)
    print(f"  Parsed: {len(intraday)} intraday, "
          f"{len(short_term)} short-term, {len(long_term)} long-term")

    conn = get_db()

    # ── Intraday verification ─────────────────────────────────
    if intraday:
        print(f"\n  Verifying intraday trades …")
        stats = _verify_intraday(conn, intraday)
        print(f"  ✓ Verified: {stats['verified']}  |  "
              f"Corrected: {stats['corrected']}  |  "
              f"Inserted: {stats['inserted']}"
              + (f"  |  Sides fixed: {stats['sides_fixed']}"
                 if stats.get('sides_fixed') else ""))

        # ── Update JSON / TXT trading reports ─────────────────
        print(f"\n  Updating trading reports …")
        report_stats = _update_trading_reports(intraday, conn)
        if report_stats["updated"]:
            print(f"  ✓ Reports updated for: {', '.join(report_stats['updated'])}")
        if report_stats["skipped"]:
            print(f"  ⊘ No report files for: {', '.join(report_stats['skipped'])}")

    # ── Capital gains import ──────────────────────────────────
    if short_term:
        n = _import_capital_gains(conn, short_term, "short_term")
        print(f"  ✓ Short-term: {n} trade(s) imported ({len(short_term) - n} already existed)")
    if long_term:
        n = _import_capital_gains(conn, long_term, "long_term")
        print(f"  ✓ Long-term:  {n} trade(s) imported ({len(long_term) - n} already existed)")

    conn.close()
    print(f"\n  Done.\n")


if __name__ == "__main__":
    main()
