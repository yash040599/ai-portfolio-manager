"""
Import and verify trade data from a Zerodha Tax P&L xlsx report.

• Intraday section  → verifies / corrects existing rows in intraday_tax_ledger,
                      inserts missing trades, marks everything as 'verified'.
• Short-term / Long-term sections → inserts into capital_gains_ledger.
• Also updates trading_data_*.json and trading_report_*.txt with corrected
  values, adds a verified header and an "Updated on ..." remark.

Usage
-----
    python scripts/shared/import_zerodha_taxpnl.py                              # all xlsx sheets in data/ZerodhaTaxPL/
    python scripts/shared/import_zerodha_taxpnl.py data/ZerodhaTaxPL/file.xlsx  # specific file
    python scripts/shared/import_zerodha_taxpnl.py --fy 2025                    # FY 2025-26 sheet only
    python scripts/shared/import_zerodha_taxpnl.py --fy 2026                    # FY 2026-27 sheet only
"""

import argparse
import datetime
import glob
import json
import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, PROJECT_ROOT)

from shared.tax_db import get_db
from config import now_ist

ZERODHA_DIR = os.path.join(PROJECT_ROOT, "data", "ZerodhaTaxPL")


# ── Protected-date helpers ────────────────────────────────────────
# A day is "protected" (skipped by all writers here) if:
#   1. It is today — Zerodha Tax P&L is only finalized T+1, so any
#      same-day data in the sheet is partial/stale.
#   2. Its trading_data_DD.json has "_reconciled": true — meaning a
#      human manually reconciled that day after a bug, and automated
#      overwrites from the sheet or the trades DB must not clobber it.

def _reconciled_dates() -> set[str]:
    """Scan all trading_data_*.json for _reconciled=true entries."""
    out: set[str] = set()
    base = os.path.join(PROJECT_ROOT, "reports", "trading")
    if not os.path.isdir(base):
        return out
    for root, _dirs, files in os.walk(base):
        for fn in files:
            if not (fn.startswith("trading_data_") and fn.endswith(".json")):
                continue
            p = os.path.join(root, fn)
            try:
                with open(p, "r", encoding="utf-8") as f:
                    d = json.load(f)
                if d.get("_reconciled") is True and d.get("date"):
                    out.add(d["date"])
            except (json.JSONDecodeError, OSError):
                continue
    return out


def _protected_dates() -> set[str]:
    """Dates that must never be auto-overwritten from the Zerodha sheet."""
    return _reconciled_dates() | {now_ist().strftime("%Y-%m-%d")}


# ── xlsx parsing ──────────────────────────────────────────────────

def _to_date_str(val) -> str:
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.isoformat()
    return str(val).strip() if val else ""


def _f(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def parse_xlsx(path: str):
    """
    Parse the 'Tradewise Exits' sheet.
    Returns (intraday_trades, short_term_trades, long_term_trades).
    Each trade is a dict with Zerodha fields.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # First sheet is always Tradewise Exits
    ws = wb.worksheets[0]

    intraday, short_term, long_term = [], [], []
    section = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        cell1 = vals[1] if len(vals) > 1 else None

        # Section headers
        if cell1 == "Equity - Intraday":
            section = "intraday"; continue
        if cell1 == "Equity - Short Term":
            section = "short_term"; continue
        if cell1 == "Equity - Long Term":
            section = "long_term"; continue
        if cell1 in ("Equity - Buyback", "Non Equity", "Mutual Funds",
                      "F&O", "Currency", "Commodity"):
            section = None; continue
        if cell1 == "Symbol":      # header row — skip
            continue

        if section is None or cell1 is None or not str(cell1).strip():
            continue

        t = {
            "symbol":       str(cell1).strip(),
            "isin":         str(vals[2]).strip() if vals[2] else "",
            "entry_date":   _to_date_str(vals[3]),
            "exit_date":    _to_date_str(vals[4]),
            "qty":          _f(vals[5]),
            "buy_value":    _f(vals[6]),
            "sell_value":   _f(vals[7]),
            "profit":       _f(vals[8]),
            "holding_days": int(_f(vals[9])),
            "fmv":          _f(vals[10]),
            "taxable_profit": _f(vals[11]),
            "turnover":     _f(vals[12]),
            "brokerage":    _f(vals[13]),
            "exchange_txn": _f(vals[14]) + _f(vals[15]),   # exch + IPFT
            "sebi_charges": _f(vals[16]),
            "gst":          _f(vals[17]) + _f(vals[18]) + _f(vals[19]),  # CGST+SGST+IGST
            "stamp_duty":   _f(vals[20]),
            "stt":          _f(vals[21]),
        }
        t["total_charges"] = round(
            t["brokerage"] + t["exchange_txn"] + t["sebi_charges"]
            + t["gst"] + t["stamp_duty"] + t["stt"], 4
        )

        if section == "intraday":
            intraday.append(t)
        elif section == "short_term":
            short_term.append(t)
        elif section == "long_term":
            long_term.append(t)

    wb.close()
    return intraday, short_term, long_term


# ── Intraday verification ────────────────────────────────────────

def _verify_intraday(conn, zerodha_trades: list[dict]) -> dict:
    """
    Verify / correct intraday_tax_ledger against Zerodha data.
    Groups by (date, symbol) and compares aggregate P&L.

    Protected-date policy:
      * Today — skipped entirely (sheet not finalized until T+1).
      * Reconciled dates (JSON flagged "_reconciled": true) — STILL checked:
        matching rows get sheet_verified='verified' marked, charges synced
        from Zerodha actuals. Mismatches only print a warning; we NEVER
        delete/replace reconciled rows, because the human fix is
        authoritative (e.g., pre-bug Zerodha sheet may disagree with the
        post-reconcile truth).

    Returns {"verified": n, "corrected": n, "inserted": n,
             "skipped_today": n, "reconciled_verified": n,
             "reconciled_mismatch": n}.
    """
    today_str = now_ist().strftime("%Y-%m-%d")
    reconciled = _reconciled_dates()
    protected = reconciled | {today_str}  # only used by the side-fixer loop

    # Group Zerodha by (exit_date, symbol) — skip only today
    z_groups: dict[tuple, list] = {}
    skipped_today = 0
    for t in zerodha_trades:
        if t["exit_date"] == today_str:
            skipped_today += 1
            continue
        key = (t["exit_date"], t["symbol"])
        z_groups.setdefault(key, []).append(t)

    # Only these dates are represented in the Zerodha sheet. We must NEVER
    # touch DB rows for dates the sheet doesn't cover — the sheet is
    # partial by nature (user may export FY-to-date excluding recent days).
    z_dates: set[str] = {date for (date, _sym) in z_groups.keys()}

    # Group DB by (date, symbol)
    db_rows = conn.execute(
        "SELECT * FROM intraday_tax_ledger"
    ).fetchall()
    db_groups: dict[tuple, list] = {}
    for r in db_rows:
        key = (r["date"], r["symbol"])
        db_groups.setdefault(key, []).append(r)

    stats = {
        "verified": 0, "corrected": 0, "inserted": 0,
        "skipped_today": skipped_today,
        "reconciled_verified": 0,
        "reconciled_mismatch": 0,
    }

    for key, z_trades in z_groups.items():
        date, symbol = key
        is_reconciled = date in reconciled
        z_total_pnl = sum(t["profit"] for t in z_trades)
        z_total_qty = sum(t["qty"] for t in z_trades)

        if key in db_groups:
            db_rows_g = db_groups[key]
            db_total_pnl = sum(r["gross_pnl"] for r in db_rows_g)
            db_total_qty = sum(r["qty"] for r in db_rows_g)

            if (abs(db_total_pnl - z_total_pnl) < 0.10
                    and abs(db_total_qty - z_total_qty) < 0.01):
                    # Match — update charges from Zerodha's actuals and mark verified.
                    # Previously only marked verified without correcting charge amounts.
                    # Zerodha's actual charges differ slightly from our estimates because
                    # they apply per-lot rounding independently; this syncs the ground truth.
                    z_total_buy  = sum(t["buy_value"]    for t in z_trades)
                    z_total_sell = sum(t["sell_value"]   for t in z_trades)
                    z_group_turnover = z_total_buy + z_total_sell
                    z_agg = {
                        "brokerage":    sum(t["brokerage"]    for t in z_trades),
                        "stt":          sum(t["stt"]          for t in z_trades),
                        "exchange_txn": sum(t["exchange_txn"] for t in z_trades),
                        "gst":          sum(t["gst"]          for t in z_trades),
                        "sebi_charges": sum(t["sebi_charges"] for t in z_trades),
                        "stamp_duty":   sum(t["stamp_duty"]   for t in z_trades),
                    }
                    z_total_charges = round(sum(z_agg.values()), 4)
                    verified_on = now_ist().isoformat()

                    for r in db_rows_g:
                        row_turnover = r["buy_value"] + r["sell_value"]
                        share = (
                            row_turnover / z_group_turnover
                            if z_group_turnover > 0
                            else 1.0 / len(db_rows_g)
                        )
                        new_total = round(z_total_charges * share, 4)
                        conn.execute(
                            """UPDATE intraday_tax_ledger
                               SET brokerage=?, stt=?, exchange_txn=?, gst=?,
                                   sebi_charges=?, stamp_duty=?, total_charges=?,
                                   net_pnl=?,
                                   verified='verified',
                                   sheet_verified='verified',
                                   sheet_verified_on=?
                               WHERE id=?""",
                            (
                                round(z_agg["brokerage"]    * share, 4),
                                round(z_agg["stt"]          * share, 4),
                                round(z_agg["exchange_txn"] * share, 4),
                                round(z_agg["gst"]          * share, 4),
                                round(z_agg["sebi_charges"] * share, 6),
                                round(z_agg["stamp_duty"]   * share, 4),
                                new_total,
                                round(r["gross_pnl"] - new_total, 2),
                                verified_on,
                                r["id"],
                            ),
                        )
                    if is_reconciled:
                        stats["reconciled_verified"] += len(db_rows_g)
                    else:
                        stats["verified"] += len(db_rows_g)
            else:
                if is_reconciled:
                    # Do NOT overwrite a human-reconciled day. Just warn.
                    stats["reconciled_mismatch"] += 1
                    print(f"    ! {symbol} on {date}: sheet disagrees with "
                          f"reconciled data (sheet P&L {z_total_pnl:+.2f} vs "
                          f"DB {db_total_pnl:+.2f}) — NOT overwriting "
                          f"(_reconciled=true). Review manually.")
                else:
                    # Mismatch — replace with Zerodha data
                    for r in db_rows_g:
                        conn.execute(
                            "DELETE FROM intraday_tax_ledger WHERE id=?",
                            (r["id"],),
                        )
                    for i, t in enumerate(z_trades):
                        _insert_zerodha_intraday(conn, t, i)
                    stats["corrected"] += len(z_trades)
                    print(f"    * Corrected {symbol} on {date}: "
                          f"P&L {db_total_pnl:+.2f} -> {z_total_pnl:+.2f}")
        else:
            if is_reconciled:
                # Sheet has a row that reconciled JSON doesn't — warn, don't insert.
                stats["reconciled_mismatch"] += 1
                print(f"    ! {symbol} on {date}: sheet has a trade but "
                      f"reconciled DB does not — NOT inserting "
                      f"(_reconciled=true). Review manually.")
            else:
                # Only in Zerodha — insert
                for i, t in enumerate(z_trades):
                    _insert_zerodha_intraday(conn, t, i)
                stats["inserted"] += len(z_trades)

    conn.commit()

    # ── Fix sides and entry/exit for all rows ─────────────────
    # Ensure side matches the trades table and entry/exit are set
    # from the correct perspective using the stored buy_value/sell_value.
    #
    # SCOPE RULES:
    #   * Only touch dates that are actually present in the Zerodha sheet
    #     (`z_dates`). If the user exports a sheet that stops at Apr 16,
    #     we must not rewrite Apr 17 rows just because they exist in the
    #     DB — the sheet has no opinion on them.
    #   * Also exclude protected dates (today + any _reconciled JSON).
    #     A protected date can have transient phantom rows in the trades
    #     table that would poison the (date, symbol, qty) -> side lookup.
    allowed_dates = z_dates - protected
    fixed = 0
    if not allowed_dates:
        all_ledger = []
        all_trades = []
    else:
        placeholders = ",".join("?" for _ in allowed_dates)
        params = tuple(allowed_dates)
        all_ledger = conn.execute(
            f"SELECT id, date, symbol, side, qty, entry_price, exit_price, "
            f"       buy_value, sell_value "
            f"FROM intraday_tax_ledger "
            f"WHERE date IN ({placeholders})",
            params,
        ).fetchall()

        # Build a lookup of trade sides: (date, symbol, qty) → side
        # — only from dates the sheet covers AND that aren't protected.
        all_trades = conn.execute(
            f"SELECT symbol, side, qty, date FROM trades "
            f"WHERE date IN ({placeholders})",
            params,
        ).fetchall()
    trade_sides: dict[tuple, str] = {}
    for tr in all_trades:
        key = (tr["date"], tr["symbol"], tr["qty"])
        trade_sides[key] = tr["side"]

    for row in all_ledger:
        key = (row["date"], row["symbol"], row["qty"])
        actual_side = trade_sides.get(key, "BUY")

        qty = row["qty"]
        if qty == 0:
            continue

        # Recalculate entry/exit from stored buy/sell values
        buy_per_unit  = round(row["buy_value"] / qty, 2)
        sell_per_unit = round(row["sell_value"] / qty, 2)

        if actual_side == "SELL":
            correct_entry = sell_per_unit   # we entered by selling
            correct_exit  = buy_per_unit    # we exited by buying back
        else:
            correct_entry = buy_per_unit    # we entered by buying
            correct_exit  = sell_per_unit   # we exited by selling

        if (row["side"] != actual_side
                or abs(row["entry_price"] - correct_entry) > 0.001
                or abs(row["exit_price"] - correct_exit) > 0.001):
            conn.execute(
                "UPDATE intraday_tax_ledger "
                "SET side=?, entry_price=?, exit_price=? WHERE id=?",
                (actual_side, correct_entry, correct_exit, row["id"]),
            )
            fixed += 1

    if fixed:
        conn.commit()
        stats["sides_fixed"] = fixed

    return stats


def _insert_zerodha_intraday(conn, t: dict, idx: int):
    """Insert a single Zerodha intraday trade into the DB.

    Cross-references the trades table to determine original side (BUY/SELL)
    and sets entry/exit from the bot's perspective accordingly.
    """
    date = t["exit_date"]
    qty = int(t["qty"])
    buy_price  = round(t["buy_value"] / qty, 2) if qty else 0
    sell_price = round(t["sell_value"] / qty, 2) if qty else 0
    order_id = f"ZV_{date}_{t['symbol']}_{idx}"

    # Look up original side from trades table
    side = "BUY"  # default
    entry_price, exit_price = buy_price, sell_price

    db_rows = conn.execute(
        "SELECT side, entry_price FROM trades WHERE date=? AND symbol=? AND qty=?",
        (date, t["symbol"], qty),
    ).fetchall()
    for row in db_rows:
        if row["side"] == "SELL":
            side = "SELL"
            entry_price = sell_price   # we entered by selling
            exit_price  = buy_price    # we exited by buying back
            break

    conn.execute(
        """INSERT OR REPLACE INTO intraday_tax_ledger
           (date, symbol, exchange, side, qty,
            entry_price, exit_price, entry_time, exit_time,
            exit_reason, gross_pnl,
            buy_value, sell_value, turnover,
            brokerage, stt, exchange_txn, gst,
            sebi_charges, stamp_duty, total_charges,
            net_pnl, order_id, verified, sheet_verified, sheet_verified_on)
           VALUES (?,?,?,?,?, ?,?,?,?, ?,?, ?,?,?, ?,?,?,?, ?,?,?, ?,?,?,?,?)""",
        (
            date, t["symbol"], "NSE", side, qty,
            entry_price, exit_price, "", "",
            "", round(t["profit"], 2),
            round(t["buy_value"], 2), round(t["sell_value"], 2),
            round(t["turnover"], 2),
            round(t["brokerage"], 4), round(t["stt"], 4),
            round(t["exchange_txn"], 4), round(t["gst"], 4),
            round(t["sebi_charges"], 4), round(t["stamp_duty"], 4),
            round(t["total_charges"], 4),
            round(t["profit"] - t["total_charges"], 2),
            order_id, "verified", "verified", now_ist().isoformat(),
        ),
    )


# ── Update trading reports (JSON + TXT) ──────────────────────────

def _trading_file_paths(date: datetime.date) -> tuple[str, str]:
    """Return (json_path, txt_path) for a trading date."""
    d = date
    base = os.path.join(
        PROJECT_ROOT, "reports", "trading",
        str(d.year), f"{d.month:02d}",
    )
    json_path = os.path.join(base, f"trading_data_{d.day:02d}.json")
    txt_path  = os.path.join(base, f"trading_report_{d.day:02d}.txt")
    return json_path, txt_path


def _match_positions_from_trades_db(conn, positions: list[dict], date_str: str):
    """
    Match JSON positions to the trades DB and update entry_price,
    exit_price, pnl from the reconciled DB values.
    The trades DB is the authoritative source (already reconciled
    with Zerodha live data).
    """
    db_rows = conn.execute(
        "SELECT symbol, side, qty, entry_price, exit_price, pnl "
        "FROM trades WHERE date=?",
        (date_str,),
    ).fetchall()

    # Build a list of unmatched DB rows
    unmatched_db = [dict(r) for r in db_rows]

    for pos in positions:
        if pos.get("status") != "CLOSED":
            continue

        symbol = pos["symbol"]
        side   = pos.get("side", "BUY")
        qty    = pos["qty"]

        # Find best matching DB row by (symbol, side, qty, closest entry)
        best = None
        best_diff = float("inf")
        for db_row in unmatched_db:
            if (db_row["symbol"] != symbol
                    or db_row["side"] != side
                    or db_row["qty"] != qty):
                continue
            diff = abs(db_row["entry_price"] - pos["entry_price"])
            if diff < best_diff:
                best = db_row
                best_diff = diff

        if best:
            unmatched_db.remove(best)
            pos["entry_price"] = best["entry_price"]
            pos["exit_price"]  = best["exit_price"]
            pos["pnl"]         = best["pnl"]


def _aggregate_zerodha_charges(z_trades: list[dict]) -> dict:
    """Sum Zerodha's actual charges across all trades for a date."""
    return {
        "brokerage":     round(sum(t["brokerage"]    for t in z_trades), 4),
        "stt":           round(sum(t["stt"]          for t in z_trades), 4),
        "exchange_txn":  round(sum(t["exchange_txn"] for t in z_trades), 4),
        "gst":           round(sum(t["gst"]          for t in z_trades), 4),
        "sebi_charges":  round(sum(t["sebi_charges"] for t in z_trades), 4),
        "stamp_duty":    round(sum(t["stamp_duty"]   for t in z_trades), 4),
    }


def _update_trading_reports(zerodha_trades: list[dict], conn) -> dict:
    """
    Update trading_data_*.json and trading_report_*.txt files.

    - Position values (entry/exit/pnl) come from the trades DB
      (already reconciled with Zerodha live data).
    - Charges come from Zerodha Tax P&L actuals.
    - Skips today's date (sheet data not finalized until T+1).
    - Skips any date whose JSON is flagged "_reconciled": true
      (human already fixed it — do not auto-overwrite).

    Returns {"updated": [dates], "skipped": [dates], "reconciled": [dates]}.
    """
    today_str = now_ist().strftime("%Y-%m-%d")
    reconciled = _reconciled_dates()

    # Group Zerodha trades by exit_date, excluding today and reconciled
    by_date: dict[str, list] = {}
    skipped_reconciled: list[str] = []
    for t in zerodha_trades:
        if t["exit_date"] == today_str:
            continue
        if t["exit_date"] in reconciled:
            if t["exit_date"] not in skipped_reconciled:
                skipped_reconciled.append(t["exit_date"])
            continue
        by_date.setdefault(t["exit_date"], []).append(t)

    stats: dict[str, list] = {
        "updated": [], "skipped": [], "reconciled": skipped_reconciled,
    }

    for date_str, z_trades in sorted(by_date.items()):
        d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        json_path, txt_path = _trading_file_paths(d)

        if not os.path.exists(json_path):
            stats["skipped"].append(date_str)
            continue

        # ── Load existing JSON ────────────────────────────────
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        positions = data.get("positions", [])
        old_charges = data.get("pnl", {}).get("charges", {})
        if not isinstance(old_charges, dict):
            old_charges = {}

        # ── Update positions from trades DB ───────────────────
        _match_positions_from_trades_db(conn, positions, date_str)

        # ── Recalculate P&L with actual Zerodha charges ──────
        z_charges = _aggregate_zerodha_charges(z_trades)
        total_buy = round(sum(t["buy_value"] for t in z_trades), 2)
        total_sell = round(sum(t["sell_value"] for t in z_trades), 2)
        total_turnover = round(total_buy + total_sell, 2)
        num_orders = old_charges.get("num_orders", len(z_trades) * 2)

        total_tax_and_charges = round(
            z_charges["brokerage"] + z_charges["stt"] +
            z_charges["exchange_txn"] + z_charges["gst"] +
            z_charges["sebi_charges"] + z_charges["stamp_duty"], 2
        )

        # Preserve Claude API cost from original report
        claude_api_cost = old_charges.get("claude_api_cost", 0.0)
        total_costs = round(total_tax_and_charges + claude_api_cost, 2)

        gross_pnl = round(
            sum(p.get("pnl", 0) for p in positions if p.get("status") == "CLOSED"), 2
        )
        net_profit = round(gross_pnl - total_costs, 2)

        # Tax estimate
        from config import Config
        tax_rate = Config.TAX_RATE_PCT * (1 + Config.TAX_CESS_PCT / 100) / 100
        tax_rate_pct = round(Config.TAX_RATE_PCT * (1 + Config.TAX_CESS_PCT / 100), 2)
        estimated_tax = round(net_profit * tax_rate, 2) if net_profit > 0 else 0.0
        profit_after_tax = round(net_profit - estimated_tax, 2)

        charges = {
            "total_turnover":        total_turnover,
            "buy_turnover":          total_buy,
            "sell_turnover":         total_sell,
            "num_orders":            num_orders,
            "brokerage":             round(z_charges["brokerage"], 2),
            "stt":                   round(z_charges["stt"], 2),
            "exchange_txn":          round(z_charges["exchange_txn"], 2),
            "gst":                   round(z_charges["gst"], 2),
            "sebi_charges":          z_charges["sebi_charges"],
            "stamp_duty":            round(z_charges["stamp_duty"], 2),
            "total_tax_and_charges": total_tax_and_charges,
            "claude_api_cost":       claude_api_cost,
            "total_costs":           total_costs,
            "zerodha_monthly_fyi":   old_charges.get("zerodha_monthly_fyi", 500.0),
        }

        pnl = {
            "gross_pnl":        gross_pnl,
            "charges":          charges,
            "net_profit":       net_profit,
            "is_profitable":    net_profit > 0,
            "tax_rate_pct":     tax_rate_pct,
            "estimated_tax":    estimated_tax,
            "profit_after_tax": profit_after_tax,
        }

        # ── Write updated JSON ────────────────────────────────
        now_str = now_ist().strftime("%Y-%m-%d %H:%M:%S")
        data["positions"] = positions
        data["pnl"] = pnl
        data["verified"] = True
        data["verified_on"] = now_str

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)

        # ── Regenerate TXT report ─────────────────────────────
        _write_verified_txt(
            txt_path, data, positions, charges, pnl, now_str
        )

        stats["updated"].append(date_str)
        print(f"    ok Updated reports for {date_str}")

    return stats


def _write_verified_txt(
    txt_path: str,
    data: dict,
    positions: list[dict],
    charges: dict,
    pnl: dict,
    verified_on: str,
):
    """Regenerate the .txt trading report from corrected data."""
    SEP_MAJOR = "=" * 58
    SEP_MINOR = "─" * 58
    SEP_TABLE = "─" * 86

    date_str = data["date"]
    mode_label = "DRY RUN (simulated)" if data.get("mode") == "dry_run" else "LIVE TRADING"
    session_count = data.get("sessions", 1)
    config = data.get("config", {})
    budget = config.get("budget", 0)
    market_condition = data.get("market_condition", "")
    trade_log = data.get("trade_log", [])

    from config import Config

    closed = [p for p in positions if p.get("status") == "CLOSED"]
    open_p = [p for p in positions if p.get("status") == "OPEN"]
    winners = [p for p in closed if p.get("pnl", 0) > 0]
    losers  = [p for p in closed if p.get("pnl", 0) < 0]

    with open(txt_path, "w", encoding="utf-8") as f:
        # ── Verified header ───────────────────────────────────
        f.write(f"{'=' * 58}\n")
        f.write("  VERIFIED -- Data corrected from Zerodha Tax P&L\n")
        f.write(f"  Updated on: {verified_on}\n")
        f.write(f"{'=' * 58}\n\n")

        # ── Header ────────────────────────────────────────────
        f.write(f"{SEP_MAJOR}\n")
        f.write(f"  INTRADAY TRADING REPORT — {date_str}\n")
        f.write(f"  Mode: {mode_label}\n")
        if session_count > 1:
            f.write(f"  Sessions: {session_count} (combined)\n")
        f.write(f"{SEP_MAJOR}\n\n")

        # ── Configuration ─────────────────────────────────────
        f.write("CONFIGURATION\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"Claude plan     : {config.get('claude_plan', 'PRO').upper()}\n")
        f.write(f"Budget          : Rs.{budget:,.2f} (from Zerodha funds)\n")
        f.write(f"Universe        : {config.get('universe', 'NIFTY100')}\n")
        if market_condition:
            f.write(f"Market condition: {market_condition}\n")
        f.write(f"Max positions   : {config.get('max_positions', 5)}\n")
        f.write(f"Stop-loss       : {config.get('stop_loss_pct', 1.5)}%\n")
        f.write(f"Target          : {config.get('target_pct', 2.0)}%\n")
        f.write(f"Circuit breaker : {Config.MAX_LOSS_PER_DAY_PCT}%\n\n")

        # ── Trade Summary ─────────────────────────────────────
        f.write("TRADE SUMMARY\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"Total trades    : {len(closed)}\n")
        f.write(f"Winners         : {len(winners)}\n")
        f.write(f"Losers          : {len(losers)}\n")
        f.write(f"Still open      : {len(open_p)}\n")
        if closed:
            win_rate = len(winners) / len(closed) * 100
            f.write(f"Win rate        : {win_rate:.1f}%\n")
        f.write("\n")

        # ── Trade Details Table ───────────────────────────────
        f.write("TRADE DETAILS\n")
        f.write(f"{SEP_TABLE}\n")
        f.write(
            f"{'SYMBOL':<12} {'SIDE':<6} {'QTY':>5} "
            f"{'ENTRY':>10} {'EXIT':>10} {'P&L':>12} "
            f"{'REASON':<14} {'ENTRY_T':<10} {'EXIT_T':<10}\n"
        )
        f.write(f"{SEP_TABLE}\n")

        for p in positions:
            exit_p  = f"Rs.{p['exit_price']:.2f}" if p.get("exit_price") else "—"
            pnl_val = f"Rs.{p.get('pnl', 0):+,.2f}" if p.get("exit_price") else "—"
            f.write(
                f"{p['symbol']:<12} {p['side']:<6} {p['qty']:>5} "
                f"Rs.{p['entry_price']:>9.2f} {exit_p:>10} {pnl_val:>12} "
                f"{(p.get('exit_reason') or 'OPEN'):<14} "
                f"{(p.get('entry_time') or '—'):<10} "
                f"{(p.get('exit_time') or '—'):<10}\n"
            )

        f.write("\n")

        # ── Rationales ────────────────────────────────────────
        f.write("TRADE RATIONALES\n")
        f.write(f"{SEP_MINOR}\n")
        for p in positions:
            f.write(f"  {p['symbol']}: {p.get('rationale', '—')}\n")
        f.write("\n")

        # ── P&L Breakdown ─────────────────────────────────────
        f.write(f"{SEP_MAJOR}\n")
        f.write("P&L BREAKDOWN\n")
        f.write(f"{SEP_MAJOR}\n\n")

        f.write(f"Gross P&L               : Rs.{pnl['gross_pnl']:+,.2f}\n\n")

        f.write("CHARGES & TAXES:\n")
        f.write(f"  Brokerage             : Rs.{charges['brokerage']:,.2f}\n")
        f.write(f"  STT (sell side)       : Rs.{charges['stt']:,.2f}\n")
        f.write(f"  Exchange transaction  : Rs.{charges['exchange_txn']:,.2f}\n")
        f.write(f"  GST (18%)             : Rs.{charges['gst']:,.2f}\n")
        f.write(f"  SEBI charges          : Rs.{charges['sebi_charges']:,.4f}\n")
        f.write(f"  Stamp duty (buy side) : Rs.{charges['stamp_duty']:,.2f}\n")
        f.write(f"  {'─' * 40}\n")
        f.write(f"  Total tax & charges   : Rs.{charges['total_tax_and_charges']:,.2f}\n\n")

        f.write("CLAUDE API COST:\n")
        f.write(f"  Claude API usage      : Rs.{charges['claude_api_cost']:,.2f}  (est. Rs.{Config.CLAUDE_COST_PER_CALL}/call × actual calls)\n")
        f.write(f"  {'─' * 40}\n")
        f.write(f"  Total all costs       : Rs.{charges['total_costs']:,.2f}\n\n")

        f.write(f"{'=' * 42}\n")
        f.write(f"  NET PROFIT AFTER ALL  : Rs.{pnl['net_profit']:+,.2f}\n")
        f.write(f"{'=' * 42}\n")
        profitable = "YES" if pnl["is_profitable"] else "NO"
        f.write(f"  Profitable?           : {profitable}\n")
        if budget > 0:
            returns_pct = pnl["net_profit"] / budget * 100
            f.write(f"  Day returns           : {returns_pct:+.2f}% on Rs.{budget:,.0f} budget\n")
        f.write("\n")

        # ── Estimated Income Tax ──────────────────────────────
        tax_rate_pct = pnl.get("tax_rate_pct", 0)
        estimated_tax = pnl.get("estimated_tax", 0)
        f.write("ESTIMATED INCOME TAX (speculative business income)\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"  Tax slab rate         : {Config.TAX_RATE_PCT}% + {Config.TAX_CESS_PCT}% cess = {tax_rate_pct}% effective\n")
        if pnl["net_profit"] > 0:
            f.write(f"  Estimated tax         : Rs.{estimated_tax:,.2f}\n")
            f.write(f"  Profit after tax      : Rs.{pnl['profit_after_tax']:+,.2f}\n")
        else:
            f.write("  Estimated tax         : Rs.0.00 (no tax on losses)\n")
            f.write("  Loss can be carried forward for 4 years (speculative only)\n")
        f.write("\n")

        f.write(f"  FYI: Zerodha Kite Connect subscription is Rs.{Config.ZERODHA_MONTHLY_COST:,.0f}/month (not deducted above).\n")
        f.write("  Track cumulative daily profits to ensure they cover this monthly cost.\n\n")

        # ── Turnover Details ──────────────────────────────────
        f.write("TURNOVER DETAILS\n")
        f.write(f"{SEP_MINOR}\n")
        f.write(f"  Buy turnover          : Rs.{charges['buy_turnover']:,.2f}\n")
        f.write(f"  Sell turnover         : Rs.{charges['sell_turnover']:,.2f}\n")
        f.write(f"  Total turnover        : Rs.{charges['total_turnover']:,.2f}\n")
        f.write(f"  Total orders          : {charges['num_orders']}\n\n")

        # ── Chronological Trade Log ───────────────────────────
        if trade_log:
            f.write("CHRONOLOGICAL TRADE LOG\n")
            f.write(f"{SEP_MINOR}\n")
            for entry in trade_log:
                f.write(
                    f"  [{entry['time']}] {entry['action']:<14} "
                    f"{entry['symbol']:<12} {entry['side']:<5} "
                    f"{entry['qty']:>5}  Rs.{entry['price']:>10}  "
                    f"{entry['detail']}\n"
                )
            f.write("\n")


# ── Capital gains import ─────────────────────────────────────────

def _import_capital_gains(conn, trades: list[dict], trade_type: str) -> int:
    """Insert capital-gains trades. Skips duplicates. Returns insert count."""
    inserted = 0
    for t in trades:
        qty = t["qty"]
        try:
            conn.execute(
                """INSERT INTO capital_gains_ledger
                   (trade_type, symbol, isin, entry_date, exit_date,
                    qty, buy_value, sell_value, profit,
                    period_of_holding, fair_market_value, taxable_profit,
                    turnover, brokerage, exchange_txn, sebi_charges,
                    gst, stamp_duty, stt, total_charges, verified)
                   VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?, ?,?,?,?, ?,?,?,?,?)""",
                (
                    trade_type, t["symbol"], t["isin"],
                    t["entry_date"], t["exit_date"],
                    qty, round(t["buy_value"], 2), round(t["sell_value"], 2),
                    round(t["profit"], 2),
                    t["holding_days"], round(t["fmv"], 2),
                    round(t["taxable_profit"], 2),
                    round(t["turnover"], 2),
                    round(t["brokerage"], 4), round(t["exchange_txn"], 4),
                    round(t["sebi_charges"], 4), round(t["gst"], 4),
                    round(t["stamp_duty"], 4), round(t["stt"], 4),
                    round(t["total_charges"], 4), "verified",
                ),
            )
            inserted += 1
        except Exception:
            pass  # duplicate — silently skip
    conn.commit()
    return inserted


# ── CLI ───────────────────────────────────────────────────────────

def _find_sheets_for_fy(fy_start: int) -> list[str]:
    """Find xlsx files matching a specific FY (e.g. fy_start=2025 → *-2025_2026-*.xlsx)."""
    pattern = os.path.join(ZERODHA_DIR, f"taxpnl-*-{fy_start}_{fy_start + 1}-*.xlsx")
    return sorted(glob.glob(pattern))


def _find_all_sheets() -> list[str]:
    """Find all taxpnl xlsx files in ZerodhaTaxPL directory."""
    pattern = os.path.join(ZERODHA_DIR, "taxpnl-*.xlsx")
    return sorted(glob.glob(pattern))


def _process_xlsx(xlsx_path: str):
    """Process a single xlsx file: verify intraday + import capital gains."""
    print(f"\n  Reading: {os.path.basename(xlsx_path)}")
    intraday, short_term, long_term = parse_xlsx(xlsx_path)
    print(f"  Parsed: {len(intraday)} intraday, "
          f"{len(short_term)} short-term, {len(long_term)} long-term")

    conn = get_db()

    # ── Intraday verification ─────────────────────────────────
    if intraday:
        print("\n  Verifying intraday trades ...")
        stats = _verify_intraday(conn, intraday)
        print(f"  ok Verified: {stats['verified']}  |  "
              f"Corrected: {stats['corrected']}  |  "
              f"Inserted: {stats['inserted']}"
              + (f"  |  Sides fixed: {stats['sides_fixed']}"
                 if stats.get('sides_fixed') else "")
              + (f"  |  Skipped today: {stats['skipped_today']}"
                 if stats.get('skipped_today') else "")
              + (f"  |  Reconciled verified: {stats['reconciled_verified']}"
                 if stats.get('reconciled_verified') else "")
              + (f"  |  Reconciled mismatches: {stats['reconciled_mismatch']}"
                 if stats.get('reconciled_mismatch') else ""))

        # ── Update JSON / TXT trading reports ─────────────────
        print("\n  Updating trading reports ...")
        report_stats = _update_trading_reports(intraday, conn)
        if report_stats["updated"]:
            print(f"  ok Reports updated for: {', '.join(report_stats['updated'])}")
        if report_stats.get("reconciled"):
            print(f"  -- Skipped reconciled (manual fix, not overwritten): "
                  f"{', '.join(report_stats['reconciled'])}")
        if report_stats["skipped"]:
            print(f"  -- No report files for: {', '.join(report_stats['skipped'])}")

    # ── Capital gains import ──────────────────────────────────
    if short_term:
        n = _import_capital_gains(conn, short_term, "short_term")
        print(f"  ok Short-term: {n} trade(s) imported ({len(short_term) - n} already existed)")
    if long_term:
        n = _import_capital_gains(conn, long_term, "long_term")
        print(f"  ok Long-term:  {n} trade(s) imported ({len(long_term) - n} already existed)")

    conn.close()
    print("\n  Done.\n")


def main():
    parser = argparse.ArgumentParser(
        description="Import Zerodha Tax P&L xlsx -- verify intraday & import capital gains.",
    )
    parser.add_argument(
        "xlsx", nargs="?", default=None,
        help="Path to Zerodha xlsx. Default: all files in data/ZerodhaTaxPL/.",
    )
    parser.add_argument(
        "--fy", type=int, default=None,
        help="Financial year start (e.g. 2025 for FY 2025-26). Picks the matching sheet automatically.",
    )
    args = parser.parse_args()

    if args.xlsx and args.fy:
        print("\n  Error: specify either a file path or --fy, not both.")
        return

    # ── Specific file ─────────────────────────────────────────
    if args.xlsx:
        if not os.path.exists(args.xlsx):
            print(f"\n  File not found: {args.xlsx}")
            return
        _process_xlsx(args.xlsx)
        return

    # ── Filter by FY ──────────────────────────────────────────
    if args.fy:
        sheets = _find_sheets_for_fy(args.fy)
        if not sheets:
            fy_label_str = f"FY {args.fy}-{str(args.fy + 1)[2:]}"
            print(f"\n  No {fy_label_str} sheet found in data/ZerodhaTaxPL/.")
            print(f"  Please download the {fy_label_str} Tax P&L xlsx from Zerodha Console")
            print("  (https://console.zerodha.com/reports/taxpnl) and place it in:")
            print(f"    {ZERODHA_DIR}/")
            print(f"  Expected filename pattern: taxpnl-*-{args.fy}_{args.fy + 1}-*.xlsx")
            return
        for sheet in sheets:
            _process_xlsx(sheet)
        return

    # ── No args: process all sheets ───────────────────────────
    sheets = _find_all_sheets()
    if not sheets:
        print("\n  No Zerodha xlsx files found in data/ZerodhaTaxPL/.")
        print("  Download from https://console.zerodha.com/reports/taxpnl")
        return
    print(f"\n  Found {len(sheets)} sheet(s) in data/ZerodhaTaxPL/")
    for sheet in sheets:
        _process_xlsx(sheet)


if __name__ == "__main__":
    main()
